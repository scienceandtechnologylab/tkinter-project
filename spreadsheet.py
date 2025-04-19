import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook

class SpreadsheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Simple Spreadsheet Software")
        self.root.geometry("600x400")
        
        # Initialize the spreadsheet grid
        self.cells = {}
        self.create_grid()
        
        # Add Save and Load buttons
        self.save_button = tk.Button(root, text="Save", command=self.save_to_excel)
        self.save_button.pack(side="left", padx=10)
        
        self.load_button = tk.Button(root, text="Load", command=self.load_from_excel)
        self.load_button.pack(side="left", padx=10)
        
    def create_grid(self):
        """Create a basic grid of entry widgets for the spreadsheet"""
        for row in range(10):  # Let's make a 10x10 grid
            for col in range(10):
                cell = tk.Entry(self.root, width=10)
                cell.grid(row=row, column=col)
                self.cells[(row, col)] = cell
    
    def save_to_excel(self):
        """Save the current spreadsheet data to an Excel file"""
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            sheet = wb.active
            for row in range(10):
                for col in range(10):
                    cell_value = self.cells[(row, col)].get()
                    sheet.cell(row=row+1, column=col+1, value=cell_value)
            wb.save(file_path)
            print(f"Spreadsheet saved to {file_path}")
    
    def load_from_excel(self):
        """Load data from an existing Excel file into the grid"""
        file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = load_workbook(file_path)
            sheet = wb.active
            for row in range(10):
                for col in range(10):
                    cell_value = sheet.cell(row=row+1, column=col+1).value
                    self.cells[(row, col)].delete(0, tk.END)  # Clear existing value
                    self.cells[(row, col)].insert(0, cell_value)  # Insert new value
            print(f"Spreadsheet loaded from {file_path}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SpreadsheetApp(root)
    root.mainloop()
